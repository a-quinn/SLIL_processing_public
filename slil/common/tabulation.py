# Author: Alastair Quinn 2022

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from slil.common.plotting_functions import figNames

class DiscreteResults():
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active # grab the active worksheet
        self.colDataOffset = 3
        self.rowDataOffset = 6

        self.cadNumberOffset = [
            '11525',
            '11526',
            '11527',
            '11534',
            '11535',
            '11536',
            '11537',
            '11538',
            '11539'
        ]

        self.ws.title = 'RU'
        self.createHeadings()
        self.ws = self.wb.create_sheet(title="FE")
        self.createHeadings()

    def save(self, filePath):
        self.wb.save(filePath)

    def createHeadings(self):
        em = '' # empty cells/spaces

        titles = []
        for i in range(5):
            titles += [[]]

        def addSpaces(l,n):
            for i in range(n):
                l += [em]

        addSpaces(titles[1], 14)
        titles[1] += ['Rotations']
        addSpaces(titles[2], 14)
        titles[2] += ['Scaphoid']
        addSpaces(titles[2], 17)
        titles[2] += ['Lunate']
        addSpaces(titles[2], 17)
        titles[2] += ['Scaphoid relative to Lunate']
        addSpaces(titles[2], 17)

        addSpaces(titles[3], 2)
        titles[3] += ['SLIL Strain']
        addSpaces(titles[3], 5)
        titles[3] += ['Wrist External Angles']
        addSpaces(titles[3], 5)
        titles3Ranges = [\
            'Flexion-Extension', em, em, em, em, em, \
            'Radial-Ulnar Deviation', em, em, em, em, em, \
            'Pro-Supination', em, em, em, em, em]
        for i in range(3):
            titles[3] += titles3Ranges

        titles[4] = ['CAD#', 'State']
        titles4Ranges = ['Min', 'SD', 'Neutral', 'SD', 'Max', 'SD']
        for i in range(11):
            titles[4] += titles4Ranges

        for i in range(len(titles)):
            self.ws.append(titles[i])

        #self.ws.merge_cells('F2:AF2')
        #self.ws.merge_cells('F3:N3')
        #self.ws.merge_cells('O3:W3')
        #self.ws.merge_cells('X3:AF3')
        of = 15
        self.ws.merge_cells(get_column_letter(of) + '2:' + get_column_letter(of + 53) + '2')
        self.ws.merge_cells(get_column_letter(of) + '3:' + get_column_letter(of + 17) + '3')
        self.ws.merge_cells(get_column_letter(of + 18) + '3:' + get_column_letter(of + 18 + 17) + '3')
        self.ws.merge_cells(get_column_letter(of + 18 + 18) + '3:' + get_column_letter(of + 18 + 18 + 17) + '3')

        offset = 3
        for col in range(11):
            start = get_column_letter(offset + col * 6)
            end = get_column_letter(offset + col * 6 + 5)
            self.ws.merge_cells(start + '4:' + end + '4')

        for row in range(1, 6):
            for col in range(1, 66):
                cell = self.ws[get_column_letter(col) + str(row)]
                cell.alignment = Alignment(horizontal="center")

        for i in range(1, 10):
            self.ws.append([i, figNames['normal']])
            self.ws.append([i, figNames['cut']])
            self.ws.append([i, figNames['scaffold']])
        
        self.ws['A1'] = 'Generated:'
        import datetime
        self.ws['B1'] = datetime.datetime.now()

    def insertSLILStrain(self, cadNumber, state, min, minSD, neutral, neutralSD, max, maxSD):

        stateOffset = ['normal', 'cut', 'scaffold']

        row = str( \
            self.rowDataOffset + \
            self.cadNumberOffset.index(cadNumber) * 3 + \
            stateOffset.index(state) \
        )
        col0 = self.colDataOffset

        def put(var, col1):
            col = get_column_letter(col1)
            self.ws[col + row] = var

        put(min, col0)
        put(minSD, col0 + 1)
        put(neutral, col0 + 2)
        put(neutralSD, col0 + 3)
        put(max, col0 + 4)
        put(maxSD, col0 + 5)

    def insertRotation(self, cadNumber, state, motion, axis, min, minSD, neutral, neutralSD, max, maxSD):

        stateOffset = ['normal', 'cut', 'scaffold']
        motionOffset = ['Scaphoid', 'Lunate', 'ScaRel2Lun']
        axisOffset = ['FE', 'RU', 'PS']

        row = str( \
            self.rowDataOffset + \
            self.cadNumberOffset.index(cadNumber) * 3 + \
            stateOffset.index(state) \
        )
        col0 = self.colDataOffset + 12 + \
            motionOffset.index(motion) * 18 + \
            axisOffset.index(axis) * 6

        def put(var, col1):
            col = get_column_letter(col1)
            self.ws[col + row] = var
            
        put(min, col0)
        put(minSD, col0 + 1)
        put(neutral, col0 + 2)
        put(neutralSD, col0 + 3)
        put(max, col0 + 4)
        put(maxSD, col0 + 5)
    
    def insertWristAngles(self, cadNumber, state, min, minSD, neutral, neutralSD, max, maxSD):

        stateOffset = ['normal', 'cut', 'scaffold']

        row = str( \
            self.rowDataOffset + \
            self.cadNumberOffset.index(cadNumber) * 3 + \
            stateOffset.index(state) \
        )
        col0 = self.colDataOffset + 6
        
        def put(var, col1):
            col = get_column_letter(col1)
            self.ws[col + row] = var
            
        put(min, col0)
        put(minSD, col0 + 1)
        put(neutral, col0 + 2)
        put(neutralSD, col0 + 3)
        put(max, col0 + 4)
        put(maxSD, col0 + 5)

    def generate(self, experiments):
        import numpy as np
        from copy import deepcopy

        def meanSTD(x):
            mean = np.mean(x)
            std = np.std(x)
            return u"{:.2f}".format(mean), u"{:.2f}".format(std)
            return u"{:.2f}\u00B1{:.2f}".format(mean, std)

        def getMean(x):
            mean = np.mean(x)
            return float(u"{:.2f}".format(mean))
        def getSTD(x):
            std = np.std(x)
            return float(u"{:.2f}".format(std))
        
        def findBoneKinematics(modelInfo, dataIn):
            motionType = dataIn[0]['type']
            self.ws = self.wb['RU']
            # raw rot data is in positive direction for both FE and RU
            # RU rotation was clock wise first
            rot = deepcopy(dataIn[0]['rot'])
            if (motionType == 'FE'):
                self.ws = self.wb['FE']
                rot *= -1.0
            # graphs are represented as righ hand so flip left hand data
            if ((motionType == 'UR') and (not modelInfo['isLeftHand'])):
                rot *= -1.0
            #time = dataIn[0]['kinematics']['time'].to_numpy()

            from slil.plotting.plots import plotKinematicsBreakDownSD_RelativeToMet_Generate
            # outputs: 9 rotation angles, 2 trials, n time points
            allX, allArrs2, allArrs = plotKinematicsBreakDownSD_RelativeToMet_Generate(dataIn, modelInfo)

            # create new arrays for scaphoid realtieve to lunate
            for indType, typeName in enumerate(allArrs2):
                tempArr = []
                for i in range(allArrs.shape[2]):
                    tempArr.append([])
                allArrs2[typeName]['ScaRelLun_rotX'] = deepcopy(tempArr)
                allArrs2[typeName]['ScaRelLun_rotY'] = deepcopy(tempArr)
                allArrs2[typeName]['ScaRelLun_rotZ'] = deepcopy(tempArr)

                for indCycle in range(allArrs.shape[2]):
                    allArrs2[typeName]['ScaRelLun_rotX'][indCycle] = \
                        allArrs2[typeName]['sca_xrot'][indCycle] - \
                        allArrs2[typeName]['lunate_flexion'][indCycle]
                    allArrs2[typeName]['ScaRelLun_rotY'][indCycle] = \
                        allArrs2[typeName]['sca_yrot'][indCycle] - \
                        allArrs2[typeName]['lunate_deviation'][indCycle]
                    allArrs2[typeName]['ScaRelLun_rotZ'][indCycle] = \
                        allArrs2[typeName]['sca_zrot'][indCycle] - \
                        allArrs2[typeName]['lunate_rotation'][indCycle]

            rotationNames = [
                'lunate_flexion', 'lunate_deviation', 'lunate_rotation',
                'sca_xrot', 'sca_yrot', 'sca_zrot',
                'hand_flexion', 'hand_deviation', 'hand_rotation',
                'ScaRelLun_rotX', 'ScaRelLun_rotY', 'ScaRelLun_rotZ']
            variableNames = ['min', 'max', 'neutral']
            discreteVars = {'normal': {}, 'cut': {}, 'scaffold': {}}
            for ind, expType in enumerate(discreteVars):
                for rotationName in rotationNames:
                    discreteVars[expType][rotationName] = {}
                    for variableName in variableNames:
                        tempArr = []
                        for i in range(allArrs.shape[2]):
                            tempArr.append(np.nan)
                        discreteVars[expType][rotationName][variableName] = deepcopy(tempArr) # what a bad way to create this...

            for expType in ['normal', 'cut', 'scaffold']:
                for ind, rotationName in enumerate(rotationNames):
                    for indCycle in range(allArrs.shape[2]):
                        #mean = np.mean(allArrs2[expType][rotationName][:])
                        #std = np.std(allArrs2[expType][rotationName][:])
                        x = allX[expType][indCycle]
                        y = allArrs2[expType][rotationName][indCycle]

                        
                        discreteVars[expType][rotationName]['min'][indCycle] = np.nanmin(y)
                        discreteVars[expType][rotationName]['max'][indCycle] = np.nanmax(y)
                        #i = np.where(np.isclose(x,0))[0][0]
                        i = (np.abs(x)).argmin()
                        try:
                            discreteVars[expType][rotationName]['neutral'][indCycle] = y[i]
                        except:
                            print('Error...')
                        #ynew = np.empty((len(xnew)), dtype=float)
                        #ynew[:] = np.nan
                        ##ynew[:] = 0.0
                        #for ind4, c in enumerate(xnew):
                        #    indexGrouped = np.where((arrX > c) & (arrX < inc + c))[0]
                        #    if len(indexGrouped) > 0:
                        #        ynew[ind4] = np.mean(y[indexGrouped])

            #for expType in ['normal', 'cut', 'scaffold']:
            #    for ind, rotationName in enumerate(rotationNames):
            #        mean = np.mean(discreteVars[expType][rotationName]['min'])
            #        std = np.std(discreteVars[expType][rotationName]['min'])
            #        print('{} {} {}+-{}'.format(expType,rotationName, mean,std))

            axis = ['FE', 'RU', 'PS']

            def setRotations(expType, bone, rotationName, rotationAcronym):
                t = discreteVars[expType][rotationName]
                self.insertRotation(modelInfo['experimentID'], expType, bone, rotationAcronym, \
                    getMean(t['min']), getSTD(t['min']), \
                    getMean(t['neutral']), getSTD(t['neutral']), \
                    getMean(t['max']), getSTD(t['max']))

            for expType in ['normal', 'cut', 'scaffold']:

                if 'FE' in motionType:
                    wristRotationName = 'hand_flexion'
                else:
                    wristRotationName = 'hand_deviation'
                t = discreteVars[expType][wristRotationName]
                self.insertWristAngles(modelInfo['experimentID'], expType, \
                    getMean(t['min']), getSTD(t['min']), \
                    getMean(t['neutral']), getSTD(t['neutral']), \
                    getMean(t['max']), getSTD(t['max']))

                setRotations(expType, 'Lunate', 'lunate_flexion', 'FE')
                setRotations(expType, 'Lunate', 'lunate_deviation', 'RU')
                setRotations(expType, 'Lunate', 'lunate_rotation', 'PS')
                setRotations(expType, 'Scaphoid', 'sca_xrot', 'FE')
                setRotations(expType, 'Scaphoid', 'sca_yrot', 'RU')
                setRotations(expType, 'Scaphoid', 'sca_zrot', 'PS')
                setRotations(expType, 'ScaRel2Lun', 'ScaRelLun_rotX', 'FE')
                setRotations(expType, 'ScaRel2Lun', 'ScaRelLun_rotY', 'RU')
                setRotations(expType, 'ScaRel2Lun', 'ScaRelLun_rotZ', 'PS')

                #rotationName = 'lunate_flexion'
                #t = discreteVars[expType][rotationName]
                #dr.insertRotation(modelInfo['experimentID'], expType, 'Lunate', 'FE', \
                #    meanSTD(t['min']), meanSTD(t['neutral']), meanSTD(t['max']))
            
        def findStrains(modelInfo, dataIn):
            motionType = dataIn[0]['type']
            self.ws = self.wb['RU']
            if (motionType == 'FE'):
                self.ws = self.wb['FE']

            from slil.common.plotting_functions import generateStrains
            strains = generateStrains(modelInfo, motionType)
            cl = 4165
            of = 2083
            len1 = len(strains[0][0])
            strainsCrop = [[],[],[]]
            strainsRemap = [1, 1, 0, 0, 2, 2] # strains order out of generateStrains() are (cut, normal, scaffold)

            from math import floor
            for ind, strain in enumerate(strains):
                for i in range(floor((len1-of)/cl)):
                    j = i+1
                    cl1 = of + (cl*i)
                    cl2 = of + (cl*j)
                    tmp = np.empty((strain.shape[0], cl2 - cl1))
                    for i in range(strain.shape[0]):
                        tmp[i,:] = strain[i][cl1:cl2]
                    strainsCrop[strainsRemap[ind]].append(tmp)


            from slil.plotting.plots import plotKinematics_Generate_cropped
            arrN, arrC, arrS = plotKinematics_Generate_cropped(modelInfo, dataIn)
            
            def flipChronologicalOrder(arrN, arrC, arrS):
                # this is a fix, otherwise they are the wrong way around
                for ind, arr in enumerate(arrN):
                    arrN[ind] = np.flip(arr, axis=0)
                for ind, arr in enumerate(arrC):
                    arrC[ind] = np.flip(arr, axis=0)
                for ind, arr in enumerate(arrS):
                    arrS[ind] = np.flip(arr, axis=0)
                return arrN, arrC, arrS
            arrN, arrC, arrS = flipChronologicalOrder(arrN, arrC, arrS)

            allArrs = np.array([arrN, arrC, arrS])
            
            strainSLIL = {'normal': {}, 'cut': {}, 'scaffold': {}}
            for ind, expType in enumerate(strainSLIL):
                tempArr = []
                for i in range(allArrs.shape[2]):
                    tempArr.append([])
                strainSLIL[expType] = deepcopy(tempArr)

            # needed incase some cycles reach different ranges
            allX = {'normal': {}, 'cut': {}, 'scaffold': {}}
            for ind, expType in enumerate(allX):
                tempArr = []
                for i in range(allArrs.shape[2]):
                    tempArr.append([])
                allX[expType] = deepcopy(tempArr) # what a bad way to create this...

            # make relative to 3rd metacarpal range of motion
            for indType, expType in enumerate(strainSLIL):
                for indCycle in range(allArrs.shape[2]):
                    if ('FE' in motionType):
                        arrX = allArrs[indType, 6, indCycle, :]
                    else:
                        arrX = allArrs[indType, 7, indCycle, :]
                    
                    inc = 0.1 # split data into groups on this sized angle
                    minX = round(float(min(arrX)) / inc) * inc
                    maxX = round(float(max(arrX)) / inc) * inc

                    xnew = np.array(np.arange(minX, maxX, inc))
                    allX[expType][indCycle] = xnew

                    ynew = np.empty((len(xnew)), dtype=float)
                    ynew[:] = np.nan
                    #ynew[:] = 0.0
                    for frameInd, c in enumerate(xnew):
                        indexGrouped = np.where((arrX > c) & (arrX < inc + c))[0]
                        if len(indexGrouped) > 0:
                            ynew[frameInd] = np.mean(strainsCrop[indType][indCycle][0][indexGrouped])
                            #ynew[frameInd] = np.mean(y[indexGrouped])

                    strainSLIL[expType][indCycle] = ynew # these are the now resampled strains
            
            discreteVarsStrain = {'normal': {}, 'cut': {}, 'scaffold': {}}
            variableNames = ['min', 'max', 'neutral']
            for ind, expType in enumerate(discreteVarsStrain):
                for variableName in variableNames:
                    tempArr = []
                    for i in range(allArrs.shape[2]):
                        tempArr.append(np.nan)
                    discreteVarsStrain[expType][variableName] = deepcopy(tempArr) # what a bad way to create this...

            for expType in ['normal', 'cut', 'scaffold']:
                for indCycle in range(allArrs.shape[2]):
                    #mean = np.mean(allArrs2[expType][rotationName][:])
                    #std = np.std(allArrs2[expType][rotationName][:])
                    x = allX[expType][indCycle]
                    y = strainSLIL[expType][indCycle]

                    # for each cycle find the minimum, maximum and (value at neutral)
                    discreteVarsStrain[expType]['min'][indCycle] = np.nanmin(y)
                    discreteVarsStrain[expType]['max'][indCycle] = np.nanmax(y)
                    #i = np.where(np.isclose(x,0))[0][0]
                    i = (np.abs(x)).argmin()
                    try:
                        discreteVarsStrain[expType]['neutral'][indCycle] = y[i]
                    except:
                        print('Error...')
            
            strainsOrdered = [[],[],[]]
            for ind, strain in enumerate(strains):
                strainsOrdered[strainsRemap[ind]] = (strain[0])
            for strain in strainsOrdered:
                strain = np.hstack(np.array(strain))
                
            strainVariance = { 'normal': 0, 'cut': 0, 'scaffold': 0 }
            strainMean = { 'normal': 0, 'cut': 0, 'scaffold': 0 }
            for ind, strain in enumerate(strainVariance):
                strainVariance[strain] = np.var(strainsOrdered[ind], axis=0)
                strainMean[strain] = np.mean(strainsOrdered[ind], axis=0)

            for expType in ['normal', 'cut', 'scaffold']:
                t = discreteVarsStrain[expType]
                self.insertSLILStrain(modelInfo['experimentID'], expType, \
                    getMean(t['min']), getSTD(t['min']), \
                    getMean(t['neutral']), getSTD(t['neutral']), \
                    getMean(t['max']), getSTD(t['max']))


        from slil.common.cache import loadOnlyExps
        from slil.common.cache import deleteCache
        models = loadOnlyExps(experiments)

        for modelInfo in models:
            cacheToDelete = [
                ['strains_', '_FE'],
                ['strains_', '_UR'],
                ]
            for cache in cacheToDelete:
                deleteCache(cache[0] + modelInfo['experimentID'] + cache[1])

            dataUR = modelInfo['dataUR']
            findBoneKinematics(modelInfo, dataUR)
            findStrains(modelInfo, dataUR)
            dataFE = modelInfo['dataFE']
            findBoneKinematics(modelInfo, dataFE)
            findStrains(modelInfo, dataFE)

        print('done')

    def format(self, motionType):
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Border, Side
        self.ws = self.wb[motionType]
        
        self.ws.column_dimensions[get_column_letter(2)].width = 10.0

        rule = ColorScaleRule(
            start_type='percentile', start_value=10, start_color='FFF8696B',
            mid_type='percentile', mid_value=50, mid_color='FFFFEB84',
            end_type='percentile', end_value=90, end_color='FF63BE7B')
        start = self.rowDataOffset
        end = 9 * 3 + self.rowDataOffset - 1
        for i in range(66):
            # whole column
            #self.ws.conditional_formatting.add(
            #    get_column_letter(3 + i) + str(start) + ':' + get_column_letter(3 + i) + str(end),
            #    rule)
            # grouped by cadaver
            for indCad in range(9):
                self.ws.conditional_formatting.add(
                    get_column_letter(3 + i) + str(start + indCad * 3) + ':' + get_column_letter(3 + i) + str(start + indCad * 3 + 2),
                    rule)
        
            self.ws.column_dimensions[get_column_letter(3 + i)].width = 6.0

            thin = Side(border_style="thin", color="000000")
            for ith in range(9):
                rowN = str(ith * 3 + 2 + self.rowDataOffset)
                for row in self.ws[get_column_letter(1 + i) + rowN +':' + get_column_letter(1 + i) + rowN]:
                    for cell in row:
                        cell.border = Border(bottom=thin)