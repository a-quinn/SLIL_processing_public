# Author: Alastair Quinn 2022

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from slil.common.plotting_functions import figNames

class DiscreteResults_Compact():
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active # grab the active worksheet
        self.colDataOffset = 3
        self.rowDataOffset = 6

        self.ws.title = 'UR'
        self.createHeadings()
        self.ws = self.wb.create_sheet(title="FE")
        self.createHeadings()

    def save(self, filePath):
        self.wb.save(filePath)

    def createHeadings(self):
        em = '' # empty cells/spaces

        titles = []
        for i in range(5):
            titles += [[]]

        def addSpaces(l,n):
            for i in range(n):
                l += [em]

        addSpaces(titles[1], 8)
        titles[1] += ['Rotations']
        addSpaces(titles[2], 8)
        titles[2] += ['Scaphoid']
        addSpaces(titles[2], 8)
        titles[2] += ['Lunate']
        addSpaces(titles[2], 8)
        titles[2] += ['Scaphoid relative to Lunate']
        addSpaces(titles[2], 8)

        addSpaces(titles[3], 2)
        titles[3] += ['SLIL Strain']
        addSpaces(titles[3], 2)
        titles[3] += ['Wrist External Angles']
        addSpaces(titles[3], 2)
        titles3Ranges = [\
            'Flexion-Extension', em, em,\
            'Radial-Ulnar Deviation', em, em,\
            'Pro-Supination', em, em]
        for i in range(3):
            titles[3] += titles3Ranges

        titles[4] = ['CAD#', 'State']
        titles4Ranges = ['Min', 'Neutral', 'Max']
        for i in range(11):
            titles[4] += titles4Ranges

        for i in range(len(titles)):
            self.ws.append(titles[i])

        #self.ws.merge_cells('F2:AF2')
        #self.ws.merge_cells('F3:N3')
        #self.ws.merge_cells('O3:W3')
        #self.ws.merge_cells('X3:AF3')
        of = 9
        self.ws.merge_cells(get_column_letter(of) + '2:' + get_column_letter(of+26) + '2')
        self.ws.merge_cells(get_column_letter(of) + '3:' + get_column_letter(of+8) + '3')
        self.ws.merge_cells(get_column_letter(of+9) + '3:' + get_column_letter(of+17) + '3')
        self.ws.merge_cells(get_column_letter(of+18) + '3:' + get_column_letter(of+26) + '3')

        offset = 3
        for col in range(10):
            start = get_column_letter(offset + col * 3)
            end = get_column_letter(offset + col * 3 + 2)
            self.ws.merge_cells(start + '4:' + end + '4')

        for row in range(1, 6):
            for col in range(1, 33):
                cell = self.ws[get_column_letter(col) + str(row)]
                cell.alignment = Alignment(horizontal="center")

        for i in range(1, 10):
            self.ws.append([i, figNames['normal']])
            self.ws.append([i, figNames['cut']])
            self.ws.append([i, figNames['scaffold']])
        
        self.ws['A1'] = 'Generated:'
        import datetime
        self.ws['B1'] = datetime.datetime.now()

    def insertSLILStrain(self, cadNumber, state, min, neutral, max):
        cadNumberOffset = [
            '11525',
            '11526',
            '11527',
            '11534',
            '11535',
            '11536',
            '11537',
            '11538',
            '11539'
        ]
        stateOffset = ['normal', 'cut', 'scaffold']

        row = str( \
            self.rowDataOffset + \
            cadNumberOffset.index(cadNumber) * 3 + \
            stateOffset.index(state) \
        )
        col0 = self.colDataOffset

        col = get_column_letter(col0)
        self.ws[col + row] = min
        col = get_column_letter(col0 + 1)
        self.ws[col + row] = neutral
        col = get_column_letter(col0 + 2)
        self.ws[col + row] = max

    def insertRotation(self, cadNumber, state, motion, axis, min, neutral, max):
        cadNumberOffset = [
            '11525',
            '11526',
            '11527',
            '11534',
            '11535',
            '11536',
            '11537',
            '11538',
            '11539'
        ]
        stateOffset = ['normal', 'cut', 'scaffold']
        motionOffset = ['Scaphoid', 'Lunate', 'ScaRel2Lun']
        axisOffset = ['FE', 'RU', 'PS']

        row = str( \
            self.rowDataOffset + \
            cadNumberOffset.index(cadNumber) * 3 + \
            stateOffset.index(state) \
        )
        col0 = self.colDataOffset + 6 + \
            motionOffset.index(motion) * 9 + \
            axisOffset.index(axis) * 3

        col = get_column_letter(col0)
        self.ws[col + row] = min
        col = get_column_letter(col0 + 1)
        self.ws[col + row] = neutral
        col = get_column_letter(col0 + 2)
        self.ws[col + row] = max
    
    def insertWristAngles(self, cadNumber, state, min, neutral, max):
        cadNumberOffset = [
            '11525',
            '11526',
            '11527',
            '11534',
            '11535',
            '11536',
            '11537',
            '11538',
            '11539'
        ]
        stateOffset = ['normal', 'cut', 'scaffold']

        row = str( \
            self.rowDataOffset + \
            cadNumberOffset.index(cadNumber) * 3 + \
            stateOffset.index(state) \
        )
        col0 = self.colDataOffset + 3

        col = get_column_letter(col0)
        self.ws[col + row] = min
        col = get_column_letter(col0 + 1)
        self.ws[col + row] = neutral
        col = get_column_letter(col0 + 2)
        self.ws[col + row] = max

    def generate(self, experiments):
        
        def run(modelInfo, dataIn):
            import numpy as np
            from copy import deepcopy
            motionType = dataIn[0]['type']
            self.ws = self.wb['UR']
            # raw rot data is in positive direction for both FE and UR
            # UR rotation was clock wise first
            rot = deepcopy(dataIn[0]['rot'])
            if (motionType == 'FE'):
                self.ws = self.wb['FE']
                rot *= -1.0
            # graphs are represented as righ hand so flip left hand data
            if ((motionType == 'UR') and (not modelInfo['isLeftHand'])):
                rot *= -1.0
            #time = dataIn[0]['kinematics']['time'].to_numpy()

            from slil.plotting.plots import plotKinematicsBreakDownSD_RelativeToMet_Generate
            # outputs: 9 rotation angles, 2 trials, n time points
            allX, allArrs2, allArrs = plotKinematicsBreakDownSD_RelativeToMet_Generate(dataIn, modelInfo)
            rotationNames = [
                'lunate_flexion', 'lunate_deviation', 'lunate_rotation',
                'sca_xrot', 'sca_yrot', 'sca_zrot',
                'hand_flexion', 'hand_deviation', 'hand_rotation']
            variableNames = ['min', 'max', 'neutral']
            discreteVars = {'normal': {}, 'cut': {}, 'scaffold': {}}
            for ind, expType in enumerate(discreteVars):
                for rotationName in rotationNames:
                    discreteVars[expType][rotationName] = {}
                    for variableName in variableNames:
                        tempArr = []
                        for i in range(allArrs.shape[2]):
                            tempArr.append(np.nan)
                        discreteVars[expType][rotationName][variableName] = deepcopy(tempArr) # what a bad way to create this...

            for expType in ['normal', 'cut', 'scaffold']:
                for ind, rotationName in enumerate(rotationNames):
                    for indCycle in range(allArrs.shape[2]):
                        #mean = np.mean(allArrs2[expType][rotationName][:])
                        #std = np.std(allArrs2[expType][rotationName][:])
                        x = allX[expType][indCycle]
                        y = allArrs2[expType][rotationName][indCycle]

                        
                        discreteVars[expType][rotationName]['min'][indCycle] = np.nanmin(y)
                        discreteVars[expType][rotationName]['max'][indCycle] = np.nanmax(y)
                        #i = np.where(np.isclose(x,0))[0][0]
                        i = (np.abs(x)).argmin()
                        try:
                            discreteVars[expType][rotationName]['neutral'][indCycle] = y[i]
                        except:
                            print('Error...')
                        #ynew = np.empty((len(xnew)), dtype=float)
                        #ynew[:] = np.nan
                        ##ynew[:] = 0.0
                        #for ind4, c in enumerate(xnew):
                        #    indexGrouped = np.where((arrX > c) & (arrX < inc + c))[0]
                        #    if len(indexGrouped) > 0:
                        #        ynew[ind4] = np.mean(y[indexGrouped])
            
            #for expType in ['normal', 'cut', 'scaffold']:
            #    for ind, rotationName in enumerate(rotationNames):
            #        mean = np.mean(discreteVars[expType][rotationName]['min'])
            #        std = np.std(discreteVars[expType][rotationName]['min'])
            #        print('{} {} {}+-{}'.format(expType,rotationName, mean,std))

            axis = ['FE', 'RU', 'PS']
            def meanSTD(x):
                mean = np.mean(x)
                std = np.std(x)
                return u"{:.2f}\u00B1{:.2f}".format(mean, std)
            def setRotations(expType, bone, rotationName, rotationAcronym):
                t = discreteVars[expType][rotationName]
                self.insertRotation(modelInfo['experimentID'], expType, bone, rotationAcronym, \
                    meanSTD(t['min']), meanSTD(t['neutral']), meanSTD(t['max']))

            for expType in ['normal', 'cut', 'scaffold']:
                self.insertSLILStrain(modelInfo['experimentID'], expType, \
                    0, 0, 0)

                if 'FE' in motionType:
                    wristRotationName = 'hand_flexion'
                else:
                    wristRotationName = 'hand_deviation'
                t = discreteVars[expType][wristRotationName]
                self.insertWristAngles(modelInfo['experimentID'], expType, \
                    meanSTD(t['min']), meanSTD(t['neutral']), meanSTD(t['max']))

                setRotations(expType, 'Lunate', 'lunate_flexion', 'FE')
                setRotations(expType, 'Lunate', 'lunate_deviation', 'RU')
                setRotations(expType, 'Lunate', 'lunate_rotation', 'PS')
                setRotations(expType, 'Scaphoid', 'sca_xrot', 'FE')
                setRotations(expType, 'Scaphoid', 'sca_yrot', 'RU')
                setRotations(expType, 'Scaphoid', 'sca_zrot', 'PS')

                #rotationName = 'lunate_flexion'
                #t = discreteVars[expType][rotationName]
                #dr.insertRotation(modelInfo['experimentID'], expType, 'Lunate', 'FE', \
                #    meanSTD(t['min']), meanSTD(t['neutral']), meanSTD(t['max']))

            print('done')

        from slil.common.cache import loadOnlyExps
        models = loadOnlyExps(experiments)
        for ind, modelInfo in enumerate(models):
            dataUR = modelInfo['dataUR']
            dataFE = modelInfo['dataFE']
            run(modelInfo, dataUR)
